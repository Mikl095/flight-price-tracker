# exporters.py
import pandas as pd
from matplotlib import pyplot as plt
from io import BytesIO
from fpdf import FPDF
import openpyxl
from utils.plotting import plot_price_history

# -----------------------------
# EXPORT CSV
# -----------------------------
def export_csv(routes, path="export.csv", route_id=None):
    data = []
    for r in routes:
        if route_id and r.get("id") != route_id:
            continue
        last_price = r.get("history")[-1]["price"] if r.get("history") else None
        min_price = min((h["price"] for h in r.get("history", [])), default=None)
        data.append({
            "id": r.get("id"),
            "origin": r.get("origin"),
            "destination": r.get("destination"),
            "departure": r.get("departure"),
            "return": r.get("return"),
            "stay_min": r.get("stay_min"),
            "stay_max": r.get("stay_max"),
            "last_price": last_price,
            "min_price": min_price,
            "target_price": r.get("target_price"),
            "notifications": r.get("notifications"),
            "email": r.get("email"),
            "travel_class": r.get("travel_class"),
            "tracking_per_day": r.get("tracking_per_day"),
        })
    df = pd.DataFrame(data)
    df.to_csv(path, index=False)
    return path

# -----------------------------
# EXPORT XLSX
# -----------------------------
def export_xlsx(routes, path="export.xlsx", route_id=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Suivis"

    headers = [
        "ID","Origin","Destination","Departure","Return","Stay Min","Stay Max",
        "Last Price","Min Price","Target Price","Notifications","Email","Class","Tracking/Day"
    ]
    ws.append(headers)

    for r in routes:
        if route_id and r.get("id") != route_id:
            continue
        last_price = r.get("history")[-1]["price"] if r.get("history") else None
        min_price = min((h["price"] for h in r.get("history", [])), default=None)
        row = [
            r.get("id"), r.get("origin"), r.get("destination"),
            r.get("departure"), r.get("return"),
            r.get("stay_min"), r.get("stay_max"),
            last_price, min_price,
            r.get("target_price"), r.get("notifications"),
            r.get("email"), r.get("travel_class"), r.get("tracking_per_day")
        ]
        ws.append(row)

    # Ajouter feuille pour graphes
    ws_graph = wb.create_sheet("Graphs")
    for r in routes:
        if route_id and r.get("id") != route_id:
            continue
        fig = plot_price_history(r.get("history", []))
        img_bytes = BytesIO()
        fig.savefig(img_bytes, format='png')
        plt.close(fig)
        img_bytes.seek(0)
        img = openpyxl.drawing.image.Image(img_bytes)
        img.anchor = 'A1'
        ws_graph.add_image(img)
    wb.save(path)
    return path

# -----------------------------
# EXPORT PDF
# -----------------------------
def export_pdf(routes, path="export.pdf", route_id=None):
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)

    for r in routes:
        if route_id and r.get("id") != route_id:
            continue
        pdf.add_page()
        pdf.set_font("Arial", 'B', 14)
        pdf.cell(0, 10, f"{r.get('origin')} → {r.get('destination')} ({r.get('travel_class')})", ln=True)
        pdf.set_font("Arial", '', 12)
        pdf.multi_cell(0, 8, f"""
ID: {r.get('id')}
Departure: {r.get('departure')}
Return: {r.get('return')}
Stay: {r.get('stay_min')}–{r.get('stay_max')} jours
Target Price: {r.get('target_price')}€
Notifications: {r.get('notifications')}
Email: {r.get('email')}
Tracking/Day: {r.get('tracking_per_day')}
""")

        # Ajouter graph
        if r.get("history"):
            fig = plot_price_history(r.get("history"))
            img_bytes = BytesIO()
            fig.savefig(img_bytes, format='png')
            plt.close(fig)
            img_bytes.seek(0)
            pdf.image(img_bytes, x=10, w=190)

    pdf.output(path)
    return path
